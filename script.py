# coding: utf8
# python 3.5.0 + openpyxl
#Ilya Fren
#  Auto mail Sender ver. 0.7
import datetime
import time
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

mail_data = []
mail_data_error = []
error_message = [0]
col_row = 0
objects = 0

ws = None
wb = None

def data_proccesing(mail_data):
    if mail_data[2] == None:
        text_log = 'Error: no name of organization in row: {}, {} in {}'.format(mail_data[1], mail_data[3],
                                                                                      mail_data[7])
        mail_data_error.append(text_log)
        print(text_log)
        return text_log
    if mail_data[3] == None:
        text_log = 'Error: no organization address in row: {}, {} in {}'.format(mail_data[1], mail_data[2],
                                                                                      mail_data[7])
        mail_data_error.append(text_log)
        print(text_log)
        return text_log
    if mail_data[4] == None:
        text_log = 'Error: no employee in row: {}, {} in {}'.format(mail_data[1], mail_data[2], mail_data[7])
        mail_data_error.append(text_log)
        print(text_log)
        return text_log
    if mail_data[5] == None:
        mail_data[5] = '-'
    if mail_data[6] == None:
        mail_data[6] = '-'
    mail_data[5] = str(mail_data[5])
    mail_data[6] = str(mail_data[6])
    auto_mail_sender(mail_data)

    return


def auto_mail_sender(data):

    #toaddr = 'www.heretic@inbox.ru'
    #toaddr = 'dispetcher@bcservice.by'
    toaddr = 'otrs@bcservice.by'
    #toaddr = 'info@bcservice.by'
    me = 'allphone@bcservice.by'
#    you = toaddr

    server = 'mail.bcservice.by'              # Сервер
    port = 2525                               # Порт
    user_name = 'allphone@bcservice.by'       # Отправитель
    user_passwd = '242425767'                 # Пароль отправителя



    #message_subject = None
    if data[0] == 1:
        message_subject = ('ТО - {}, {}'.format(data[2], data[3]))
        message_text = ('Ответственный за ТО: ' + data[4]+'\n' + 'ТО - ' + data[2] + ',   ' + data[3] + '\n' +
                    'кассы: ' + data[5] + '    весы: ' + data[6])

    if data[0] == 0:
        message_subject = 'Отчет об ошибках в списке ТО'
        message_text = ('Курганов\n' + data[1])

    # Формируем заголовок письма
    msg = MIMEMultipart('mixed')
    msg['Subject'] = message_subject
    msg['From'] = me
    msg['To'] = toaddr

    # Формируем письмо
    part = MIMEText(message_text)
    msg.attach(part)

    # Подключение
    s = smtplib.SMTP(server, port)
    s.ehlo()
    s.starttls()
    s.ehlo()
    # Авторизация
    s.login(user_name, user_passwd)
    # Отправка письма
    s.sendmail(me, toaddr, msg.as_string())
    s.quit()
    if data[0] == 1:
        text_log = 'mail to {} was sended, object: {}, Time:{}'.format(data[4], data[3], datetime.datetime.now())
    if data[0] == 0:
        text_log = 'Error mail was sended'
    print(text_log)
    return text_log

if __name__ == "__main__":

    #wb = load_workbook(filename='test.xlsx')
    wb = load_workbook(filename='Статистика по объектам ТО.xlsx')
    temp_list = wb.get_sheet_names()
    temp_list = temp_list[:(len(temp_list)-4)]
    print('City in work book: '+(' '.join(temp_list)))
    # Обработка страниц из xlsx
    for city in temp_list:
        ws = wb[city]
        i = 2
        col_row = 0
        while ws.cell(row=i, column=1).value != None:
            mail_data.append(1)
            for row in ws.iter_rows(min_row=i, min_col=1, max_col=6, max_row=i):
                for cell in row:
                    mail_data.append(cell.value)
            mail_data.append(wb[city])
            data_proccesing(mail_data)
            col_row += 1
            i += 1
            mail_data = []
            time.sleep(2)
        print('objects count in {}:{}'.format(wb[city], col_row))
        objects += col_row
        ws = None
    # Если обнаруженны ошибочные строки отправляем письмо
    if len(mail_data_error) > 0:
        error_message.append('\n'.join(mail_data_error))
        auto_mail_sender(error_message)
    wb = None
    print('objects count{}'.format(objects))