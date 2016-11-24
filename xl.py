# coding: utf8
from openpyxl import load_workbook
wb = load_workbook(filename='test.xlsx')
temp_list = wb.get_sheet_names()
temp_list = temp_list[3:(len(temp_list)-1)]
print(temp_list)
ws = wb[temp_list[0]]
print(wb[temp_list[0]])
wb = None
# for city in temp_list:
#     ws = wb[city]
#     i = 2
#     col_row = 0
#     while ws.cell(row=i, column=2).value != None:
#         for row in ws.iter_rows(min_row=i, min_col=2, max_col=6, max_row=i):
#             for cell in row:
#                 mail_data.append(cell.value)
#         print(mail_data)
#         col_row += 1
#         i += 1
#         mail_data = []
#     print('objects count in {}:{}'.format(temp_list[index], col_row))






